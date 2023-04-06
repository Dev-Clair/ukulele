# Import the Tkinter Class
import tkinter as tk
import tkinter.ttk as ttk
from tkinter import PhotoImage
from tkinter.constants import *
from tkinter.messagebox import showerror, showinfo
# Import -SQLite3- Database
import sqlite3
# Import CSV - Module
from csv import writer
# Import OPENPYXL Module
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


class Connection:
    def __init__(self, filename: str):
        """
            Create a context manager for better resource (database connection) management
        """
        self.filename = filename
        self.connection = sqlite3.connect(self.filename)

    def __enter__(self):
        self.cursor = self.connection.cursor()
        return self.cursor

    def __exit__(self, exc_typ, exc_val, exc_tb):
        if self.cursor:
            self.connection.commit()
            self.connection.close()


# Define Query Variables
# Create Table
create_table = "CREATE TABLE IF NOT EXISTS surveytable (Tag INTEGER PRIMARY KEY NOT NULL, Name TEXT NOT NULL, Age INT NOT NULL, Email TEXT, Gender TEXT NOT NULL, Ethnicity TEXT NOT NULL, Disability TEXT NOT NULL, Enjoyed TEXT, Curious TEXT, Science TEXT, Future TEXT)"
# Display Records
display_record = "SELECT * FROM surveytable"
# Select Record
select_record = "SELECT * FROM surveytable WHERE Age BETWEEN ? AND ? AND Gender=? AND Ethnicity=? AND DISABILITY=?"
# Update Record - Column Should Only be Updated as Required
update_record = "UPDATE surveytable SET Email=? WHERE Tag=?"
# Delete Record
delete_record = "DELETE * FROM surveytable WHERE Tag=?"


def createtable():
    """
        Creates table in database
    """
    with Connection('surveydb.db') as connection:
        connection.execute(create_table)


def displayrecord() -> list:
    """
        Displays all record(s) in table
    """
    with Connection('surveydb.db') as connection:
        rows = connection.execute(display_record)
        records = rows.fetchall()
        return records


def selectrecord(lower_range_value=" ", upper_range_value=" ", gender=" ", ethnicity=" ", disability=" ") -> list:
    """
        Selects record(s) from table
    """
    with Connection('surveydb.db') as connection:
        values = (lower_range_value, upper_range_value,
                  gender, ethnicity, disability)
        rows = connection.execute(
            select_record, values)
        records = rows.fetchall()
        return records


def updaterecord(column=" ", tag=" "):
    """
        Updates student record in table
    """
    with Connection('surveydb.db') as connection:
        values = (column, tag)
        connection.execute(update_record, values)


def deleterecord(tag=" "):
    """
        Deletes selected record from database
    """
    with Connection('surveydb.db') as connection:
        connection.execute(delete_record, tag)


def total_num() -> int:
    """
        returns total number of rows/records in the table
    """
    totalno = "SELECT COUNT(*) FROM surveytable"
    with Connection('surveydb.db') as connection:
        totalval1 = connection.execute(totalno)
        totalval2 = totalval1.fetchone()
        totalcount = totalval2[0]
        return totalcount


def womenpercent() -> int:
    """
        returns percentage of 'female' women in the table
    """
    women_per = "SELECT COUNT(*) FROM surveytable WHERE Gender=?"
    with Connection('surveydb.db') as connection:
        total_women1 = connection.execute(women_per, ("Female",))
        total_women2 = total_women1.fetchone()
        total_women3 = total_women2[0]
        total_women = total_women3

    totalno = "SELECT COUNT(*) FROM surveytable"
    with Connection('surveydb.db') as connection:
        totalval1 = connection.execute(totalno)
        totalval2 = totalval1.fetchone()
        totalval3 = totalval2[0]
        totalval = totalval3

    return round((total_women/totalval)*100, 2)


def enjoytour() -> int:
    """
        returns average number of people who enjoyed the tour
    """
    aver_enj = "SELECT COUNT(*) FROM surveytable WHERE Enjoyed=?"
    with Connection('surveydb.db') as connection:
        enjoy1 = connection.execute(aver_enj, ("Yes",))
        enjoy2 = enjoy1.fetchone()
        enjoy = enjoy2[0]
        return enjoy


def curious() -> int:
    """
        returns average number of people curious about the singing sculpture
    """
    aver_cur = "SELECT COUNT(*) FROM surveytable WHERE Curious=?"
    with Connection('surveydb.db') as connection:
        curious1 = connection.execute(aver_cur, ("Yes",))
        curious2 = curious1.fetchone()
        curious = curious2[0]
        return curious


def science() -> int:
    """
        returns total number of people who will like to learn more about the science of acoustics
    """
    aver_sci = "SELECT COUNT(*) FROM surveytable WHERE Science=?"
    with Connection('surveydb.db') as connection:
        science1 = connection.execute(aver_sci, ("Yes",))
        science2 = science1.fetchone()
        science = science2[0]
        return science


def future() -> int:
    """
        returns total number of people who will like to attend future events
    """
    future_eve = "SELECT COUNT(*) FROM surveytable WHERE Future=?"
    with Connection('surveydb.db') as connection:
        future1 = connection.execute(future_eve, ("Yes",))
        future2 = future1.fetchone()
        future = future2[0]
        return future


class Mainframe:
    def __init__(self, master):
        self.master = master
        self.master.title("UKULELE - Admin")
        self.master.geometry('1350x680+5+5')
        # self.master.iconimage = PhotoImage(file='pic\icon_img.png')
        # self.master.iconphoto(True, self.master.iconimage)
        self.master.minsize(100, 100)
        self.master.resizable(0, 0)

        # Widget Styling
        apply_style = ttk.Style(master)
        apply_style.configure('TFrame')
        apply_style.configure('TLabel', foreground="black",
                              font=("Century", 10, "bold"))
        apply_style.configure('TEntry', foreground="blue",
                              background="WhiteSmoke", font=("Helvetica", 12, "normal"))
        apply_style.configure('TButton', foreground="blue",
                              background="MistyRose", font=("Bell MT", 8, "normal"))
        apply_style.configure('TRadiobutton', foreground="blue",
                              font=("times new roman", 10, "bold"))
        apply_style.configure('TMenubutton', foreground="blue",
                              font=("Helvetica", 10, "bold"))
        # apply_style.configure('TLabelframe', foreground="violet red",
        #                       font=("Helvetica", 10, "bold"))
        apply_style.configure(
            'TSeparator', foreground="blue", background="blue")
        apply_style.configure(
            'Treeview.Heading', foreground="lightgreen", background="lightblue")

       # Create Title Label
        slidertext = 'UKULELE - \"The Singing Sculpture\"'  # Slider Text
        self.topLabel = ttk.Label(
            master, text=slidertext, style='TLabel')
        self.topLabel.config(font=("Bell MT", 30, "bold"))
        self.topLabel.place(x=10, y=0)
        # Show-All Button
        self.showButton = ttk.Button(
            master, text='Show All Results', cursor='hand2',  style='TButton', command=self.showall, state="disabled")
        self.showButton.place(x=1065, y=0)
        # Database Button
        self.databaseButton = ttk.Menubutton(
            master, text='Database', cursor='hand2', direction="left", style='TMenubutton')
        self.databaseButton.place(x=1165, y=0)
        # Configure Database Menubutton
        self.databaseButton.menu = tk.Menu(
            self.databaseButton, tearoff=False, cursor='hand2')
        self.databaseButton["menu"] = self.databaseButton.menu
        self.databaseButton.menu.add_command(
            label="Connect to Database", command=self.connectdb)
        self.databaseButton.menu.add_command(
            label="Upload Database to Server", command=self.uploaddb)
        # Export Button
        self.exportButton = ttk.Menubutton(
            master, text='Export', cursor='hand2', direction="left", style='TMenubutton')
        self.exportButton.place(x=1265, y=0)
        # Configure Export Menubutton
        self.exportButton.menu = tk.Menu(
            self.exportButton, tearoff=False)
        self.exportButton["menu"] = self.exportButton.menu
        self.exportButton.menu.add_command(
            label="...export to Excel", command=self.exportexcel)
        self.exportButton.menu.add_command(
            label="...export to CSV", command=self.exportcsv)
        self.exportButton.menu.add_command(
            label="...export as XML", command=self.exportxml)

        # Create TOP LEFT Frame
        topleftframe = ttk.Frame(
            master)
        topleftframe.place(x=10, y=60, width=450, height=250)

        # Create Child Widgets
        # Title Label
        self.titlelabel = ttk.Label(
            topleftframe, text="Kindly select the options you'll like to see results for:", style='TLabel', justify=LEFT)
        self.titlelabel.config(font=("times new roman", 10, "bold"))
        self.titlelabel.config(foreground="black")
        self.titlelabel.pack(side=TOP, expand=1)

        # Age
        ageframe = ttk.Labelframe(
            topleftframe, text="Age:", labelanchor=NW)
        ageframe.pack(side=TOP, expand=0, fill=X, padx=2)
        # Lower Bound
        lowerage = [i for i in range(15, 46)]
        self.loweragevar = tk.IntVar(value=lowerage[0])
        self.loweragelabel = ttk.Label(
            ageframe, text="Select Lower Age Bound: ", style="TLabel")
        self.loweragelabel.config(font=("Bell MT", 8, "bold"))
        self.loweragelabel.pack(side=LEFT, expand=1, anchor=W)
        self.loweragedrop = ttk.OptionMenu(
            ageframe, self.loweragevar, *lowerage, direction="right", style='TMenubutton')
        self.loweragedrop.pack(side=LEFT, expand=1, anchor=W)
        # Upper Bound
        upperage = [i for i in range(46, 76)]
        self.upperagevar = tk.IntVar(value=upperage[0])
        self.upperagedrop = ttk.OptionMenu(
            ageframe, self.upperagevar, *upperage, direction="right", style='TMenubutton')
        self.upperagedrop.pack(side=RIGHT, expand=1, anchor=E)
        self.upperagelabel = ttk.Label(
            ageframe, text="Select Upper Age Bound: ", style="TLabel")
        self.upperagelabel.config(font=("Bell MT", 8, "bold"))
        self.upperagelabel.pack(side=RIGHT, expand=1, anchor=E)

        # Gender
        genderframe = ttk.Labelframe(
            topleftframe, text="Gender:", labelanchor=NW)
        genderframe.pack(side=TOP, expand=0, fill=X, padx=2)
        gender = ["--Click to Select Gender--", "Male", "Female", "Transgender", "Gender-neutral", "Non-binary",
                  "Agender", "Pangender", "Genderqueer", "Two-spirit", "Third-gender"]
        self.gendervar = tk.StringVar(value=gender[0])
        self.genderdrop = ttk.OptionMenu(
            genderframe, self.gendervar, *gender, direction="right", style='TMenubutton')
        self.genderdrop.pack(side=TOP, expand=1, anchor=CENTER)

        # Ethnicity/Race
        ethnicframe = ttk.Labelframe(
            topleftframe, text="Ethnicity:", labelanchor=NW)
        ethnicframe.pack(side=TOP, expand=0, fill=X, padx=2)
        ethnic = ["--Click to Select Race--", "Arab", "Black",
                  "Chinese", "Indo-Asian", "Latin", "White"]
        self.ethnicvar = tk.StringVar(value=ethnic[0])
        self.ethnicdrop = ttk.OptionMenu(
            ethnicframe, self.ethnicvar, *ethnic, direction="right", style='TMenubutton')
        self.ethnicdrop.pack(side=TOP, expand=1, anchor=CENTER)

        # Disability
        disabilityframe = ttk.Labelframe(
            topleftframe, text="Disability:", labelanchor=NW)
        disabilityframe.pack(side=TOP, expand=0, fill=X, padx=2)
        self.disabilityvar = tk.StringVar()
        self.disability_firstradio = ttk.Radiobutton(
            disabilityframe, text='Yes', value="Yes", variable=self.disabilityvar, style='TRadiobutton')
        self.disability_firstradio.pack(
            side=LEFT, expand=1, padx=2, pady=2, anchor=W)
        self.disability_secondradio = ttk.Radiobutton(
            disabilityframe, text='No', value="No", variable=self.disabilityvar, style='TRadiobutton')
        self.disability_secondradio.pack(
            side=LEFT, expand=1, padx=2, anchor=W)

        # Create Submit Button
        self.submitbutton = ttk.Button(
            topleftframe, text="Submit", command=self.submitentry, style='TButton')
        self.submitbutton.pack(
            side=TOP, expand=1, padx=5, pady=5, anchor=CENTER)

        # Create Separator
        displayseparator = ttk.Separator(master, style='TSeparator')
        displayseparator.place(x=10, y=330, width=450)

        # Create BOTTOM LEFT Frame
        bottomleftframe = ttk.Frame(
            master)
        bottomleftframe.place(x=10, y=340, width=450, height=350)

        # Create Child Widgets
        # Title Label
        self.titlelabel = ttk.Label(
            bottomleftframe, text="QUICK SURVEY SUMMARY", style='TLabel', justify=LEFT)
        self.titlelabel.config(font=("times new roman", 10, "bold"))
        self.titlelabel.config(foreground="black")
        self.titlelabel.pack(side=TOP, expand=0)
        # Total Number based on Selection
        totalframe = ttk.Labelframe(
            bottomleftframe, text="Total No.:", labelanchor=NW)
        totalframe.pack(side=TOP, expand=0, fill=X,
                        anchor=CENTER, padx=2)
        self.totalvar = tk.IntVar()
        self.totallabel = ttk.Label(
            totalframe, textvariable=self.totalvar, justify=CENTER)
        self.totallabel.pack(side=TOP, expand=0, padx=2, pady=3)
        # Percentage of Women
        womenframe = ttk.Labelframe(
            bottomleftframe, text="Percentage (%) Of Women:", labelanchor=NW)
        womenframe.pack(side=TOP, expand=0, fill=X,
                        anchor=CENTER, padx=2)
        self.womenvar = tk.DoubleVar()
        self.womenlabel = ttk.Label(
            womenframe, textvariable=self.womenvar, justify=CENTER)
        self.womenlabel.pack(side=TOP, expand=0, padx=2, pady=3)
        # Average No. Who Enjoyed the Tour
        enjoyframe = ttk.Labelframe(
            bottomleftframe, text="Average No. Of People Who Enjoyed The Tour:", labelanchor=NW)
        enjoyframe.pack(side=TOP, expand=0, fill=X,
                        anchor=CENTER, padx=2)
        self.enjoyvar = tk.IntVar()
        self.enjoylabel = ttk.Label(
            enjoyframe, textvariable=self.enjoyvar, justify=CENTER)
        self.enjoylabel.pack(side=TOP, expand=0, padx=5, pady=3)
        # Average No. Curious about the Sculpture
        curiousframe = ttk.Labelframe(
            bottomleftframe, text="Average No. Of People Curious About The Ukulele:", labelanchor=NW)
        curiousframe.pack(side=TOP, expand=0, fill=X,
                          anchor=CENTER, padx=2)
        self.curiousvar = tk.IntVar()
        self.curiouslabel = ttk.Label(
            curiousframe, textvariable=self.curiousvar, justify=CENTER)
        self.curiouslabel.pack(side=TOP, expand=0, padx=2, pady=3)
        # Average No. Who will like to learn more about the Science of Acoustics
        scienceframe = ttk.Labelframe(
            bottomleftframe, text="Average No. Who Will Like To Learn More About The Science Of Acoustics:", labelanchor=NW)
        scienceframe.pack(side=TOP, expand=0, fill=X,
                          anchor=CENTER, padx=2)
        self.sciencevar = tk.IntVar()
        self.sciencelabel = ttk.Label(
            scienceframe, textvariable=self.sciencevar, justify=CENTER)
        self.sciencelabel.pack(side=TOP, expand=0, padx=2, pady=3)
        # No. Who will like to attend future events
        futureframe = ttk.Labelframe(
            bottomleftframe, text="No. Of People Who Will Like To Attend Future Events:", labelanchor=NW)
        futureframe.pack(side=TOP, expand=0, fill=X,
                         anchor=CENTER, padx=2)
        self.futurevar = tk.IntVar()
        self.futurelabel = ttk.Label(
            futureframe, textvariable=self.futurevar, justify=CENTER)
        self.futurelabel.pack(side=TOP, expand=0, padx=2, pady=3)

        # Create RIGHT Frame
        rightframe = ttk.Frame(master)
        # Create Vertical and Horizontal Scrollbars
        self.hsbar = ttk.Scrollbar(
            rightframe, orient="horizontal")
        self.vsbar = ttk.Scrollbar(
            rightframe)
        # Create Treeview Widget
        self.treecolumn = ("Tag", "Name", "Age", "Email", "Gender",
                           "Ethnicity", "Disability", "Enjoyed", "Curious", "Science", "Future")
        self.table = ttk.Treeview(
            rightframe, columns=self.treecolumn, show="headings", style="Treewview.Heading", xscrollcommand=self.hsbar.set, yscrollcommand=self.vsbar.set, cursor='hand2', selectmode=EXTENDED)
        # Configure Scrollbars to -treeview- Table
        self.hsbar.config(command=self.table.xview)
        self.vsbar.config(command=self.table.yview)
        # Define Table -Treeview- Headings
        self.table.heading("Tag", text="Tag No.",
                           anchor=CENTER)  # Add Callback
        self.table.heading("Name", text="Name", anchor=CENTER)  # Add Callback
        self.table.heading("Age", text="Age", anchor=CENTER)  # Add Callback
        self.table.heading("Email", text="Email",
                           anchor=CENTER)  # Add Callback
        self.table.heading("Gender", text="Gender",
                           anchor=CENTER)  # Add Callback
        self.table.heading("Ethnicity", text="Ethnicity",
                           anchor=CENTER)  # Add Callback
        self.table.heading("Disability", text="Disability",
                           anchor=CENTER)  # Add Callback
        self.table.heading("Enjoyed", text="Enjoyed",
                           anchor=CENTER)  # Add Callback
        self.table.heading(
            "Curious", text="Curious", anchor=CENTER)  # Add Callback
        self.table.heading(
            "Science", text="Science", anchor=CENTER)  # Add Callback
        self.table.heading("Future", text="Future",
                           anchor=CENTER)  # Add Callback
        # Configure Table -Treeview- Columns
        self.table.column("Tag", width=75, stretch=1, anchor=CENTER)
        self.table.column("Name", width=200, stretch=1, anchor=CENTER)
        self.table.column("Age", width=100, stretch=1, anchor=CENTER)
        self.table.column("Email", width=200, stretch=1, anchor=CENTER)
        self.table.column("Gender", width=100, stretch=1, anchor=CENTER)
        self.table.column("Ethnicity", width=100, stretch=1, anchor=CENTER)
        self.table.column("Disability", width=100, stretch=1, anchor=CENTER)
        self.table.column("Enjoyed", width=100, stretch=1, anchor=CENTER)
        self.table.column("Curious", width=100, stretch=1, anchor=CENTER)
        self.table.column("Science", width=150, stretch=1, anchor=CENTER)
        self.table.column("Future", width=100, stretch=1, anchor=CENTER)
        # Display -Pack- RIGHT Frame Widgets
        self.hsbar.pack(side=BOTTOM, fill=X)
        self.vsbar.pack(side=RIGHT, fill=Y)
        self.table.pack(fill=BOTH, expand=1)
        rightframe.place(x=470, y=60, width=880, height=600)

        # Copyright Label
        self.cpyrightLabel = ttk.Label(
            master, text="© Copyright 2023 | Lexi-Clair Designs", style='TLabel')
        self.cpyrightLabel.config(font=("times new roman", 8, "bold"))
        self.cpyrightLabel.config(foreground="blue")
        self.cpyrightLabel.config(background="light grey")
        self.cpyrightLabel.place(x=0, y=665, width=1350)

    def showall(self: object):
        """
            On-click, shows table on treeview after connection to database
        """
        for item in self.table.get_children():
            self.table.delete(item)
        self.datarecord = displayrecord()
        for record in sorted(self.datarecord, key=lambda record: record[0]):
            self.table.insert(parent="", index=END, values=record)
        # BOTTOM LEFT FRAME Label FUnction Call
        self.totalnumber()
        self.percentwomen()
        self.averageenjoyed()
        self.averagecurious()
        self.averagescience()
        self.numberfuture()
        self.totalnumber()
        self.percentwomen()
        self.averageenjoyed()
        self.averagecurious()
        self.averagescience()
        self.numberfuture()

    def connectdb(self: object):
        """
            connects to database and displays table on treeview
        """
        try:
            createtable()
            showinfo('Database', 'Database Connection Successful')
            self.showButton.config(state="normal")
        except:
            showerror('Database', 'Database Connection was Unsuccessful')
            self.showButton.config(state="disabled")

    def uploaddb(self: object):
        """
            uploads database to server
        """
        print("Filezilla will hold database for client")

    def exportexcel(self: object):
        """
            exports database -all or selection- to Ms-Excel
        """
        self.excelexport = tk.Toplevel()
        self.excelexport.title('Select Export')
        self.excelexport.geometry("250x70+450+250")
        self.excelexport.resizable(0, 0)

        self.excelchoicevar = tk.StringVar()
        self.exportdatabase = ttk.Radiobutton(self.excelexport, text="Export Database to Excel (.xlsx)",
                                              value="db.xlsx", variable=self.excelchoicevar, style='TRadiobutton', cursor='hand2', command=self.excelexportselect)
        self.exportdatabase.pack(side=TOP, fill=X, padx=5, pady=5)
        self.exporttreeview = ttk.Radiobutton(self.excelexport, text="Export Treeview to Excel (.xlsx)",
                                              value="tree.xlsx", variable=self.excelchoicevar, style='TRadiobutton', cursor='hand2', command=self.excelexportselect)
        self.exporttreeview.pack(side=TOP, fill=X, padx=5, pady=5)

    def excelexportselect(self: object):
        """
            credits: https://medium.com/@azmideliaslan/how-to-output-excel-with-openpyxl-python-c1039436dc24
        """
        if self.excelchoicevar.get() == "db.xlsx":
            # Create Workbook
            self.wb = Workbook()
            # Get Active Sheet
            self.ws = self.wb.active
            # Create Headings
            self.ws.cell(row=1, column=1).value = "Tag"
            self.ws.cell(row=1, column=2).value = "Name"
            self.ws.cell(row=1, column=3).value = "Age"
            self.ws.cell(row=1, column=4).value = "Email"
            self.ws.cell(row=1, column=5).value = "Gender"
            self.ws.cell(row=1, column=6).value = "Ethnicity"
            self.ws.cell(row=1, column=7).value = "Disability"
            self.ws.cell(row=1, column=8).value = "Enjoyed"
            self.ws.cell(row=1, column=9).value = "Curious"
            self.ws.cell(row=1, column=10).value = "Science"
            self.ws.cell(row=1, column=11).value = "Future"
            # Style Headings
            for val in range(1, 12):
                self.ws.cell(row=1, column=val).font = Font(
                    bold=True, name='Times New Roman', size=12)
                self.ws.cell(row=1, column=val).fill = PatternFill(
                    patternType="solid", fgColor="00e5ee")
                self.ws.cell(row=1, column=val).alignment = Alignment(
                    horizontal="center")
            # Set Column Dimensions- Width
            self.ws.column_dimensions["A"].width = 7
            self.ws.column_dimensions["B"].width = 28
            self.ws.column_dimensions["D"].width = 38
            for col in "CEFGHIJK":
                self.ws.column_dimensions[col].width = 15
            # Retrieve Data from Database
            self.databaseimport = displayrecord()
            # Insert into Excel
            for self.ws_row, row in enumerate(self.databaseimport):
                for self.ws_col, item in enumerate(row):
                    self.ws.cell(row=self.ws_row + 2,
                                 column=self.ws_col + 1).value = item
                    self.ws.cell(row=self.ws_row + 2, column=self.ws_col +
                                 1).alignment = Alignment(horizontal="center")

            self.wb.save("db_survey_report.xlsx")

        if self.excelchoicevar.get() == "tree.xlsx":
            # Create Workbook
            self.wb = Workbook()
            # Get Active Sheet
            self.ws = self.wb.active
            # Create Headings
            self.ws.cell(row=1, column=1).value = "Tag"
            self.ws.cell(row=1, column=2).value = "Name"
            self.ws.cell(row=1, column=3).value = "Age"
            self.ws.cell(row=1, column=4).value = "Email"
            self.ws.cell(row=1, column=5).value = "Gender"
            self.ws.cell(row=1, column=6).value = "Ethnicity"
            self.ws.cell(row=1, column=7).value = "Disability"
            self.ws.cell(row=1, column=8).value = "Enjoyed"
            self.ws.cell(row=1, column=9).value = "Curious"
            self.ws.cell(row=1, column=10).value = "Science"
            self.ws.cell(row=1, column=11).value = "Future"
            # Style Headings
            for val in range(1, 12):
                self.ws.cell(row=1, column=val).font = Font(
                    bold=True, name='Times New Roman', size=12)
                self.ws.cell(row=1, column=val).fill = PatternFill(
                    patternType="solid", fgColor="00e5ee")
                self.ws.cell(row=1, column=val).alignment = Alignment(
                    horizontal="center")
            # Set Column Dimensions- Width
            self.ws.column_dimensions["A"].width = 7
            self.ws.column_dimensions["B"].width = 28
            self.ws.column_dimensions["D"].width = 38
            for col in "CEFGHIJK":
                self.ws.column_dimensions[col].width = 15
            # Retrieve Displayed Treeview Data from Database
            self.treeviewexport = self.viewrecord
            # Insert into Excel
            for self.ws_row, row in enumerate(self.treeviewexport):
                for self.ws_col, item in enumerate(row):
                    self.ws.cell(row=self.ws_row + 2,
                                 column=self.ws_col + 1).value = item
                    self.ws.cell(row=self.ws_row + 2, column=self.ws_col +
                                 1).alignment = Alignment(horizontal="center")

            self.wb.save("tree_survey_report.xlsx")

    def exportcsv(self: object):
        """
            exports database -all or selection- to CSV file
        """
        self.csvexport = tk.Toplevel()
        self.csvexport.title('Select Export')
        self.csvexport.geometry("250x70+450+250")
        self.csvexport.resizable(0, 0)

        self.csvchoicevar = tk.StringVar()
        self.exportdatabase = ttk.Radiobutton(self.csvexport, text="Export Database to CSV (.csv)",
                                              value="db.csv", variable=self.csvchoicevar, style='TRadiobutton', cursor='hand2', command=self.csvexportselect)
        self.exportdatabase.pack(side=TOP, fill=X, padx=5, pady=5)
        self.exporttreeview = ttk.Radiobutton(self.csvexport, text="Export Treeview to CSV (.csv)",
                                              value="tree.csv", variable=self.csvchoicevar, style='TRadiobutton', cursor='hand2', command=self.csvexportselect)
        self.exporttreeview.pack(side=TOP, fill=X, padx=5, pady=5)

    def csvexportselect(self: object):
        if self.csvchoicevar.get() == "db.csv":
            headings = ["Tag", "Name", "Age", "Email", "Gender",
                        "Ethnicity", "Disability", "Enjoyed", "Curious", "Science", "Future"]
            with open('new_survey_report.csv', 'w', newline='', encoding="UTF-8") as self.newfile:
                # Create Writer object
                self.write = writer(self.newfile)
                # Write Headings
                self.write.writerow(headings)
                # Write Rows
                self.write.writerows(self.datarecord)

        if self.csvchoicevar.get() == "tree.csv":
            headings = ["Tag", "Name", "Age", "Email", "Gender",
                        "Ethnicity", "Disability", "Enjoyed", "Curious", "Science", "Future"]
            with open('tree_survey_report.csv', 'w', newline='', encoding="UTF-8") as self.newfile:
                # Create Writer object
                self.write = writer(self.newfile)
                # Write Headings
                self.write.writerow(headings)
                # Write Rows
                self.write.writerows(self.viewrecord)

    def exportxml(self: object):
        """
            exports table -all or selection- to XML
        """
        showerror("Error", "This feature is currently unavailable")

    def submitentry(self: object):
        # Displays Result For Selected Options
        for item in self.table.get_children():
            self.table.delete(item)
        self.viewrecord = selectrecord(self.loweragevar.get(), self.upperagevar.get(), self.gendervar.get(),
                                       self.ethnicvar.get(), self.disabilityvar.get())
        for record in sorted(self.viewrecord, key=lambda record: record[0]):
            self.table.insert(parent="", index=END, values=record)

    # Result Functions - To be derived from Treeview Display
    def totalnumber(self: object) -> int:
        """
            Calculates the total number of attendees based on selection and
            updates the self.totalvar variable afterwards
        """
        self.value = total_num()
        self.totalvar.set(value=self.value)

    def percentwomen(self: object) -> int:
        """
            Calculates the percentage of women based on selection and
            updates the self.womenvar variable afterwards
        """
        self.value = womenpercent()
        self.womenvar.set(value=self.value)

    def averageenjoyed(self: object) -> int:
        """
            Calculates the average number of attendees based on selection who enjoyed the tour and
            updates the self.enjoyvar variable afterwards
        """
        self.value = enjoytour()
        self.enjoyvar.set(value=self.value)

    def averagecurious(self: object) -> int:
        """
            Calculates the average number of attendees based on selection who are curious as to how the sculpture sings and
            updates the self.curiousvar variable afterwards
        """
        self.value = curious()
        self.curiousvar.set(value=self.value)

    def averagescience(self: object) -> int:
        """
            Calculates the average number of attendees based on selection who will like to learn more about science and
            updates the self.sciencevar  afterwards
        """
        self.value = science()
        self.sciencevar.set(value=self.value)

    def numberfuture(self: object) -> int:
        """
            Calculates the number of attendees based on selection who will like to attend future
            and updates the self.futurevar variable afterwards
        """
        self.value = future()
        self.futurevar.set(value=self.value)


if __name__ == "__main__":
    root = tk.Tk()
    window = Mainframe(root)
    root.mainloop()
